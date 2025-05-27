import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl import Workbook 
import numpy as np


class ExcelReader():

    def __init__(self) -> None:
        self.DataFrame = pd.DataFrame({})
        
    """
    Read data in the file
    - param FilePath: the file path
    return: data in the file (Dataframe structure in pandas)
    """
    def LoadFile(self, FilePath):
        if re.match('^.*?\.csv$',FilePath):
            self.FileType = 'csv'
            self.DataFrame = pd.read_csv(FilePath)
            #print('This is a CSV file')
        elif re.match('^.*?\.(xls|xlsx)$',FilePath):
            self.FileType = 'excel'
            self.DataFrame = pd.read_excel(FilePath)
            #print('This is a xls file')
        else:
            self.FileType = 'error'
            print('Unsupported file type !')
    
 
    def DeriveLine2Csv(self, DerivePath, Header, Data):
        self.DataFrame = {Header, Data}
        self.DataFrame.to_csv(DerivePath, index=False)


class ExcelWriter():

    def __init__(self, ExcelPath, ExcelName) -> None:
        self.excel_path = ExcelPath
        self.excel_name = ExcelName
        self.excel_full_path = ExcelPath + ExcelName
        self.excel_is_null = 0
        if os.path.exists(self.excel_full_path):
            self.LoadExcel()
            self.excel_is_null += self.ExcelIsNull()
        else:
            self.CreateExcel()

    def LoadExcel(self):
        self.workbook = load_workbook(self.excel_full_path)
        self.sheet = self.workbook.worksheets[0]

    def CreateExcel(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active  
        self.workbook.save(filename= self.excel_path + self.excel_name)

    def ExcelIsNull(self):
        
        for row in self.sheet.iter_rows():
            # Any non-empty cell in the file will consider that the Excel is not empty. 
            if any(cell.value for cell in row):
                return 1
        return 0

    def AddRowData2Excel(self, Data):

        max_row_num = self.sheet.max_row
        self.sheet._current_row = max_row_num
        for i in range(Data.shape[1]):
            self.sheet.cell(max_row_num + self.excel_is_null, i+1).value = Data[0][i]
        self.workbook.save(self.excel_full_path)

    def AddLineData2Excel(self, Data):

        max_col_num = self.sheet.max_column
        for i in range(Data.shape[0]):
            # The initial coordinate in openpyxl is (1,1), but not (0,0)
            self.sheet.cell(i+1, max_col_num + self.excel_is_null).value = Data[i][0] 
        self.workbook.save(self.excel_full_path)
    
    
    """
    Add Data to Excel
    - param Data: The data needs to store (The dimension should be [x,1] or [1,x])
    return: None
    """
    def AddData2Excel(self, Data):
        self.excel_is_null = self.ExcelIsNull()
        if len(Data.shape) != 2:
            print('Dimension Error !')
        
        if Data.shape[0] == 1:
            self.AddRowData2Excel(Data)
        elif Data.shape[1] == 1:
            self.AddLineData2Excel(Data)
        else:
            print("Dimension ERROR !")

        
